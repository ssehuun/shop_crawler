#-*- coding: utf-8 -*-
import re
import warnings
import pprint
import shutil
import subprocess
import time
import urllib
import json
import requests
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
warnings.filterwarnings('ignore')

client_id = "RD1oPcv8ybcFOPSXm2xm"
client_secret = "asZCz3L4qr"

top500_dic = dict()
review_cnt_list = []
apiitem_list = []
filename = []

def getOnePageList(v):
    onePageitemList = []
    cookies = {
        'NNB': 'RN7EMFYA62RWE',
        'ASID': '798dc239000001867ec5ac6800000058',
        'SHP_BUCKET_ID': '2',
        'nid_inf': '-1107154688',
        'NID_AUT': 'AoQvnWYFT9S3uVGVXO+RyUUMWU6O2pwLmb5xhzW1TkjboQB47aAywEDe2aA88LaA',
        'NID_JKL': 'TpDt+CMfZfv6bFsbVEbo4i0HlEk47/CkxQFaTsmDSU8=',
        'nx_ssl': '2',
        'page_uid': 'iuQS+sp0Jy0ssju/siCssssssIV-021446',
        'spage_uid': 'iuQS%2Bsp0Jy0ssju%2FsiCssssssIV-021446',
        'ncpa': '6844584|lfxlqjeg|599ad2aad5b68b19991edbacafcde7ee2ed1b6a2|s_30aec8d3505be|bfd10bc1d9f035c55aead7b4158621780bf7806c:6877236|lfxlvntc|0fb397515c9c3c62f22a231e1278818834e4a9aa|s_c4e2e0c4754a|1d388fa3f48c24bccce033ca51a9e4a30b401da4:1236626|lfxma3vk|e198145465569e0e596a1c954af8f98427d71be5|s_15bd4c7c1dcb5|e0f8c8eb9443b7c95d15cf4aa6107cdb2d2c54ba:823793|lfxmhsvs|fd070c31106b7eed98ffbff6ca9fbdea14f89495|s_3020ce80b43ba|0e36a1cd82ad97f6f17aabb7cdf87cfc25350227:1127474|lfxmjrvc|6011eb4ac2dce31670cb97f379c20cc73f868488|s_22c3f940d7252|4d01b48472d663e05216ce2837c5f606afe30a7a:6278965|lfxmtqo0|e6e133c47946a266f209ecd98bb5057f8a8a8337|s_1955b1ce015c4|e9db3ec85851cd16bebe793fbc85efdc0b1899b2',
        'NID_SES': 'AAABpeBd4euK/6Ac4lpx8vAP0b6T58aJlsc8ywl7IQmWDC3ppDhdZKhZCkqIK0Ij7I/jr5ZPzL+/ZR+dPf1CdsGPAVPgY7yXL3Q90malAQWnt2ugkv0p16rG3RUWrAdZFcdgdqZzoQxvWR7BFbBnBTpy5tHAoytR50TiVTscceT8EAKVffahFxSjnTjDjKbWi88y0Fsd8ET9yZPFnQl/gb3lzYYHskBhGmIjnjFenafkBdmpaLVMktkGZf2qF928akqH23e2Neou3kSD+TGnUK3vOU6xwZQAO8w6jgT/TCcQYNLfgznnGAVLCDi2xXW7Z/pNBKHAssRZDdTZXaK3E+kK6NGZaHu8MR0R3JigfZz8757lvt/6n0ITIjNGKYFAiuDkvwoEOj7Gw7Dt9/Y7PUJ4+UB7nY/PXUEi3t/5Ye+5H9XCfKCzCvBaAiSoOY9PpiVDboYjOjjDOuW2r7Rawj0GkO26DZELA41keafVdE80N89Sh6UbHF+1Mm1sDHKxITwio0YKMJBIYdPW/2vlAUNXrJ/f4CTfZVqBCV5rPT5F4+KAMntbqaJaYfAbJg7dL3UEXg==',
    }

    headers = {
        'authority': 'search.shopping.naver.com',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'cookie': 'NNB=RN7EMFYA62RWE; ASID=798dc239000001867ec5ac6800000058; SHP_BUCKET_ID=2; nid_inf=-1107154688; NID_AUT=AoQvnWYFT9S3uVGVXO+RyUUMWU6O2pwLmb5xhzW1TkjboQB47aAywEDe2aA88LaA; NID_JKL=TpDt+CMfZfv6bFsbVEbo4i0HlEk47/CkxQFaTsmDSU8=; nx_ssl=2; page_uid=iuQS+sp0Jy0ssju/siCssssssIV-021446; spage_uid=iuQS%2Bsp0Jy0ssju%2FsiCssssssIV-021446; ncpa=6844584|lfxlqjeg|599ad2aad5b68b19991edbacafcde7ee2ed1b6a2|s_30aec8d3505be|bfd10bc1d9f035c55aead7b4158621780bf7806c:6877236|lfxlvntc|0fb397515c9c3c62f22a231e1278818834e4a9aa|s_c4e2e0c4754a|1d388fa3f48c24bccce033ca51a9e4a30b401da4:1236626|lfxma3vk|e198145465569e0e596a1c954af8f98427d71be5|s_15bd4c7c1dcb5|e0f8c8eb9443b7c95d15cf4aa6107cdb2d2c54ba:823793|lfxmhsvs|fd070c31106b7eed98ffbff6ca9fbdea14f89495|s_3020ce80b43ba|0e36a1cd82ad97f6f17aabb7cdf87cfc25350227:1127474|lfxmjrvc|6011eb4ac2dce31670cb97f379c20cc73f868488|s_22c3f940d7252|4d01b48472d663e05216ce2837c5f606afe30a7a:6278965|lfxmtqo0|e6e133c47946a266f209ecd98bb5057f8a8a8337|s_1955b1ce015c4|e9db3ec85851cd16bebe793fbc85efdc0b1899b2; NID_SES=AAABpeBd4euK/6Ac4lpx8vAP0b6T58aJlsc8ywl7IQmWDC3ppDhdZKhZCkqIK0Ij7I/jr5ZPzL+/ZR+dPf1CdsGPAVPgY7yXL3Q90malAQWnt2ugkv0p16rG3RUWrAdZFcdgdqZzoQxvWR7BFbBnBTpy5tHAoytR50TiVTscceT8EAKVffahFxSjnTjDjKbWi88y0Fsd8ET9yZPFnQl/gb3lzYYHskBhGmIjnjFenafkBdmpaLVMktkGZf2qF928akqH23e2Neou3kSD+TGnUK3vOU6xwZQAO8w6jgT/TCcQYNLfgznnGAVLCDi2xXW7Z/pNBKHAssRZDdTZXaK3E+kK6NGZaHu8MR0R3JigfZz8757lvt/6n0ITIjNGKYFAiuDkvwoEOj7Gw7Dt9/Y7PUJ4+UB7nY/PXUEi3t/5Ye+5H9XCfKCzCvBaAiSoOY9PpiVDboYjOjjDOuW2r7Rawj0GkO26DZELA41keafVdE80N89Sh6UbHF+1Mm1sDHKxITwio0YKMJBIYdPW/2vlAUNXrJ/f4CTfZVqBCV5rPT5F4+KAMntbqaJaYfAbJg7dL3UEXg==',
        'logic': 'PART',
        'referer': 'https://search.shopping.naver.com/search/all?query=%ED%8A%B8%EC%9C%84%EB%93%9C%EC%9E%90%EC%BC%93&prevQuery=%ED%8A%B8%EC%9C%84%EB%93%9C%EC%9E%90%EC%BC%93',
        'sbth': '4231bedc931199611fa09b2600b4c94496fe7e44ca75d8a25488851acb770a4b2e5d19d862b00fb86a2c89e0aec4edca',
        'sec-ch-ua': '"Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
    }

    params = {
        'deliveryFee': '',
        'deliveryTypeValue': '',
        'eq': '',
        'iq': '',
        'origQuery': v,
        'pagingIndex': '1',
        'pagingSize': '40',
        'productSet': 'total',
        'query': v,
        'sort': 'rel',
        'viewType': 'list',
        'xq': '',
    }
    response = requests.get('https://search.shopping.naver.com/api/search/all', params=params, cookies=cookies, headers=headers)
    itemsObj = json.loads(response.text)
    itemlist = itemsObj['shoppingResult']['products']
    for i in itemlist:
        productName = i['productName']
        price = i['price'] # lowPrice
        openDate = i['openDate']
        opDate = openDate[0:4]+'.'+openDate[4:6]
        reviewCount = i['reviewCount']
        purchaseCnt = i['purchaseCnt']
        mallProductUrl = i['mallProductUrl']
        onePageitemList.append(productName)
        onePageitemList.append(price)
        onePageitemList.append(opDate)
        onePageitemList.append(reviewCount)
        onePageitemList.append(purchaseCnt)
        onePageitemList.append(mallProductUrl)
        
    return onePageitemList
    # print(productName, price, opDate, reviewCount, purchaseCnt, mallProductUrl)

def searchKeyword():    
    for k, v in top500_dic.items():
        print(k, v)
        encText = urllib.parse.quote(v)
        url = "https://openapi.naver.com/v1/search/shop?query=" + encText + "&display=40"
        request = urllib.request.Request(url)
        request.add_header("X-Naver-Client-Id", client_id)
        request.add_header("X-Naver-Client-Secret", client_secret)
        response = urllib.request.urlopen(request)
        rescode = response.getcode()
        if(rescode==200):
            res_body = response.read()
            json_obj = json.loads(res_body.decode('utf-8'))
            apiitem_list = json_obj["items"]
            print(apiitem_list)
        else:
            print("Error Code:" + rescode)
        return


def getData():
    wb = Workbook()
    for k, v in top500_dic.items():
        try:
            ws = wb.create_sheet(v)
        except ValueError:
            ws = wb.create_sheet(v.split('/')[0])
        if 'Sheet' in wb.sheetnames:
            wb.remove_sheet(wb['Sheet'])
        ws.append((['상품명', '가격', '등록일', '리뷰', '구매건수', '스토어링크']))
        # li = getOnePageList(v)
        
        cookies = {
            'NNB': 'RN7EMFYA62RWE',
            'ASID': '798dc239000001867ec5ac6800000058',
            'SHP_BUCKET_ID': '2',
            'nid_inf': '-1107154688',
            'NID_AUT': 'AoQvnWYFT9S3uVGVXO+RyUUMWU6O2pwLmb5xhzW1TkjboQB47aAywEDe2aA88LaA',
            'NID_JKL': 'TpDt+CMfZfv6bFsbVEbo4i0HlEk47/CkxQFaTsmDSU8=',
            'nx_ssl': '2',
            'page_uid': 'iuQS+sp0Jy0ssju/siCssssssIV-021446',
            'spage_uid': 'iuQS%2Bsp0Jy0ssju%2FsiCssssssIV-021446',
            'ncpa': '6844584|lfxlqjeg|599ad2aad5b68b19991edbacafcde7ee2ed1b6a2|s_30aec8d3505be|bfd10bc1d9f035c55aead7b4158621780bf7806c:6877236|lfxlvntc|0fb397515c9c3c62f22a231e1278818834e4a9aa|s_c4e2e0c4754a|1d388fa3f48c24bccce033ca51a9e4a30b401da4:1236626|lfxma3vk|e198145465569e0e596a1c954af8f98427d71be5|s_15bd4c7c1dcb5|e0f8c8eb9443b7c95d15cf4aa6107cdb2d2c54ba:823793|lfxmhsvs|fd070c31106b7eed98ffbff6ca9fbdea14f89495|s_3020ce80b43ba|0e36a1cd82ad97f6f17aabb7cdf87cfc25350227:1127474|lfxmjrvc|6011eb4ac2dce31670cb97f379c20cc73f868488|s_22c3f940d7252|4d01b48472d663e05216ce2837c5f606afe30a7a:6278965|lfxmtqo0|e6e133c47946a266f209ecd98bb5057f8a8a8337|s_1955b1ce015c4|e9db3ec85851cd16bebe793fbc85efdc0b1899b2',
            'NID_SES': 'AAABpeBd4euK/6Ac4lpx8vAP0b6T58aJlsc8ywl7IQmWDC3ppDhdZKhZCkqIK0Ij7I/jr5ZPzL+/ZR+dPf1CdsGPAVPgY7yXL3Q90malAQWnt2ugkv0p16rG3RUWrAdZFcdgdqZzoQxvWR7BFbBnBTpy5tHAoytR50TiVTscceT8EAKVffahFxSjnTjDjKbWi88y0Fsd8ET9yZPFnQl/gb3lzYYHskBhGmIjnjFenafkBdmpaLVMktkGZf2qF928akqH23e2Neou3kSD+TGnUK3vOU6xwZQAO8w6jgT/TCcQYNLfgznnGAVLCDi2xXW7Z/pNBKHAssRZDdTZXaK3E+kK6NGZaHu8MR0R3JigfZz8757lvt/6n0ITIjNGKYFAiuDkvwoEOj7Gw7Dt9/Y7PUJ4+UB7nY/PXUEi3t/5Ye+5H9XCfKCzCvBaAiSoOY9PpiVDboYjOjjDOuW2r7Rawj0GkO26DZELA41keafVdE80N89Sh6UbHF+1Mm1sDHKxITwio0YKMJBIYdPW/2vlAUNXrJ/f4CTfZVqBCV5rPT5F4+KAMntbqaJaYfAbJg7dL3UEXg==',
        }

        headers = {
            'authority': 'search.shopping.naver.com',
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            # 'cookie': 'NNB=RN7EMFYA62RWE; ASID=798dc239000001867ec5ac6800000058; SHP_BUCKET_ID=2; nid_inf=-1107154688; NID_AUT=AoQvnWYFT9S3uVGVXO+RyUUMWU6O2pwLmb5xhzW1TkjboQB47aAywEDe2aA88LaA; NID_JKL=TpDt+CMfZfv6bFsbVEbo4i0HlEk47/CkxQFaTsmDSU8=; nx_ssl=2; page_uid=iuQS+sp0Jy0ssju/siCssssssIV-021446; spage_uid=iuQS%2Bsp0Jy0ssju%2FsiCssssssIV-021446; ncpa=6844584|lfxlqjeg|599ad2aad5b68b19991edbacafcde7ee2ed1b6a2|s_30aec8d3505be|bfd10bc1d9f035c55aead7b4158621780bf7806c:6877236|lfxlvntc|0fb397515c9c3c62f22a231e1278818834e4a9aa|s_c4e2e0c4754a|1d388fa3f48c24bccce033ca51a9e4a30b401da4:1236626|lfxma3vk|e198145465569e0e596a1c954af8f98427d71be5|s_15bd4c7c1dcb5|e0f8c8eb9443b7c95d15cf4aa6107cdb2d2c54ba:823793|lfxmhsvs|fd070c31106b7eed98ffbff6ca9fbdea14f89495|s_3020ce80b43ba|0e36a1cd82ad97f6f17aabb7cdf87cfc25350227:1127474|lfxmjrvc|6011eb4ac2dce31670cb97f379c20cc73f868488|s_22c3f940d7252|4d01b48472d663e05216ce2837c5f606afe30a7a:6278965|lfxmtqo0|e6e133c47946a266f209ecd98bb5057f8a8a8337|s_1955b1ce015c4|e9db3ec85851cd16bebe793fbc85efdc0b1899b2; NID_SES=AAABpeBd4euK/6Ac4lpx8vAP0b6T58aJlsc8ywl7IQmWDC3ppDhdZKhZCkqIK0Ij7I/jr5ZPzL+/ZR+dPf1CdsGPAVPgY7yXL3Q90malAQWnt2ugkv0p16rG3RUWrAdZFcdgdqZzoQxvWR7BFbBnBTpy5tHAoytR50TiVTscceT8EAKVffahFxSjnTjDjKbWi88y0Fsd8ET9yZPFnQl/gb3lzYYHskBhGmIjnjFenafkBdmpaLVMktkGZf2qF928akqH23e2Neou3kSD+TGnUK3vOU6xwZQAO8w6jgT/TCcQYNLfgznnGAVLCDi2xXW7Z/pNBKHAssRZDdTZXaK3E+kK6NGZaHu8MR0R3JigfZz8757lvt/6n0ITIjNGKYFAiuDkvwoEOj7Gw7Dt9/Y7PUJ4+UB7nY/PXUEi3t/5Ye+5H9XCfKCzCvBaAiSoOY9PpiVDboYjOjjDOuW2r7Rawj0GkO26DZELA41keafVdE80N89Sh6UbHF+1Mm1sDHKxITwio0YKMJBIYdPW/2vlAUNXrJ/f4CTfZVqBCV5rPT5F4+KAMntbqaJaYfAbJg7dL3UEXg==',
            'logic': 'PART',
            'referer': 'https://search.shopping.naver.com/search/all?query=%ED%8A%B8%EC%9C%84%EB%93%9C%EC%9E%90%EC%BC%93&prevQuery=%ED%8A%B8%EC%9C%84%EB%93%9C%EC%9E%90%EC%BC%93',
            'sbth': '4231bedc931199611fa09b2600b4c94496fe7e44ca75d8a25488851acb770a4b2e5d19d862b00fb86a2c89e0aec4edca',
            'sec-ch-ua': '"Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
        }

        params = {
            'deliveryFee': '',
            'deliveryTypeValue': '',
            'eq': '',
            'iq': '',
            'origQuery': v,
            'pagingIndex': '1',
            'pagingSize': '40',
            'productSet': 'total',
            'query': v,
            'sort': 'rel',
            'viewType': 'list',
            'xq': '',
        }
        response = requests.get('https://search.shopping.naver.com/api/search/all', params=params, cookies=cookies, headers=headers)
        itemsObj = json.loads(response.text)
        itemlist = itemsObj['shoppingResult']['products']
        
        for i in itemlist:
            oneItem= []
            productTitle = i['productTitle']
            price = i['price'] # lowPrice
            openDate = i['openDate']
            opDate = openDate[0:4]+'.'+openDate[4:6]
            reviewCount = i['reviewCount']
            purchaseCnt = i['purchaseCnt']
            mallProductUrl = i['mallProductUrl']
            oneItem.append(productTitle)
            oneItem.append(price)
            oneItem.append(opDate)
            oneItem.append(reviewCount)
            oneItem.append(purchaseCnt)
            oneItem.append(mallProductUrl)
            ws.append(oneItem)
        print(k, v, "완료")
    global filename
    filename = '_'.join(filename)
    filename = re.sub("[/ ]", "", filename)
    wb.save('C:\\Users\\LG\\.virtualenvs\\'+filename+'.xlsx')
    wb.close()
        # pprint.pprint(item_list)
    return

        
def writeExcel():
    wb = Workbook()
    for i in top500_dic:
        ws = wb.create_sheet(top500_dic[i])
        if 'Sheet' in wb.sheetnames:
            wb.remove_sheet(wb['Sheet'])
        ws.append((['상품명', '가격', '등록일', '리뷰', '구매건수', '스토어링크']))
        for i in item_list:
            ws.append(i)
    wb.save('C:\\Users\\user\\.virtualenvs\\naverdatalabtop500.xlsx')
    wb.close()


def getRankKeyword():
    options = Options()
    options.add_argument("headless")
    # options.add_experimental_option("excludeSwitches", ["enable-logging"]) #크롬 종료되지 않는 상태 유지

    options.binary_location= 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe'

    driver = webdriver.Chrome(executable_path='C:\\Users\\user\\Downloads\\chromedriver_win32\\chromedriver.exe', chrome_options = options) ## 크롬 드라이버가 위치한 경로 대입 필요
    driver.get('https://datalab.naver.com/shoppingInsight/sCategory.naver')

    time.sleep(1)

    # driver.find_element(By.XPATH, '//*[@id="18_device_0"]').click() #기기별 전체 클릭
    # time.sleep(0.5)
    # driver.find_element(By.XPATH, '//*[@id="19_gender_0"]').click() #성별 전체 클릭
    # time.sleep(0.5)
    # driver.find_element(By.XPATH, '//*[@id="20_age_0"]').click() #연령별 전체 클릭
    # time.sleep(0.5)

    driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/span').click() #분야 클릭
    time.sleep(0.5)

    filename.append(driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/ul/li[5]/a').text)
    # driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/ul/li[9]/a').click() #패션의류 클릭, 마지막 분야
    driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/ul/li[5]/a').click() #가구/인테리어
    time.sleep(0.5)
    #//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/ul/li[9]/a #생활건강

    
    driver.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/span').click() # 2분류 클릭
    time.sleep(0.5)

    filename.append(driver.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li[9]/a').text)
    # driver.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li[8]/a').click() # 여성의류 클릭
    driver.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li[9]/a').click() # 인테리어소품
    time.sleep(0.5)
    # //*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li[8]/a #반려동물
    

    # driver.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[3]/span').click() # 3분류 클릭
    # time.sleep(0.5)

    # driver.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[3]/ul/li[1]/a').click() # 니트/스웨터 클릭
    # time.sleep(0.5)

    driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/a').click() #조회하기 클릭
    time.sleep(1)

    for i in range(0, 12) : # 25페이지까지
        for j in range(1, 21) : # 20개씩
            path = f'//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{j}]/a'
            result = driver.find_element(By.XPATH, path).text
            top500_dic[result.split('\n')[0]] = result.split('\n')[1]
            
            # time.sleep(0.1)
            # ws.append(result.split('\n'))
        
        driver.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/a[2]').click()
        time.sleep(0.1)
    driver.close()
    driver.quit()
    print(top500_dic, "네이버 데이터랩 Top 500 키워드 뽑기 완료")

# searchKeyword()

# writeExcel()
if __name__ == '__main__':
    start = time.time()
    getRankKeyword()
    getData()
    end = time.time()
    print("총시간: ", end-start)